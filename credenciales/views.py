import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from .models import Credencial  # Asegúrate de usar el nombre exacto de tu modelo

# 1. Vista para la pantalla de bienvenida limpia (Inicio)
def inicio(request):
    return render(request, 'credenciales/inicio.html')

# 2. Vista de registro (formulario y tabla)
def registro_credenciales(request):
    if request.method == "POST":
        cedula = request.POST.get('cedula')
        apellidos_nombres = request.POST.get('apellidos_nombres')
        turno = request.POST.get('turno')
        
        # Guarda el registro en la base de datos
        Credencial.objects.create(
            cedula=cedula,
            apellidos_nombres=apellidos_nombres,
            turno=turno
        )
        return redirect('registro_credenciales')

    registros = Credencial.objects.all()
    return render(request, 'credenciales/credenciales.html', {'registros': registros})

# 3. Vista para eliminar todos los registros
def limpiar_registros(request):
    if request.method == "POST":
        Credencial.objects.all().delete()
    return redirect('registro_credenciales')

# 4. Vista para exportar los datos a Excel
def exportar_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="cuadrilla_registros.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"
    
    # Encabezados de las columnas
    columns = ['N°', 'Cédula', 'Apellidos y Nombres', 'Turno']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        
    # Llenar con los datos de la base de datos
    registros = Credencial.objects.all()
    for row_num, reg in enumerate(registros, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=reg.cedula)
        ws.cell(row=row_num, column=3, value=reg.apellidos_nombres)
        ws.cell(row=row_num, column=4, value=reg.turno)
        
    wb.save(response)
    return response